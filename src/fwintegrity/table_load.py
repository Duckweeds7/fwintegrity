"""Explicit table sources + column mapping → canonical change rows (`change_type`, `source`, …)."""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from importlib import resources
from pathlib import Path
from typing import Any, TypeAlias

from .ticket import slugify_column_name


@dataclass(frozen=True, slots=True)
class ChangeRowMapping:
    """Map file / dict column labels to canonical row keys used by parsers."""

    change_kind: str
    source: str
    destination: str
    service: str
    ticket_number: str | None = None
    inf_number: str | None = None


AUDIT_EXPORT_DEFAULT_MAPPING = ChangeRowMapping(
    change_kind="Change Type",
    source="Source",
    destination="Destination",
    service="Service",
)

TICKET_CSV_DEFAULT_MAPPING = ChangeRowMapping(
    change_kind="Action",
    source="Source IP Address",
    destination="Destination IP Address",
    service="Service Port",
    ticket_number="Ticket Number",
    inf_number="INF Number",
)


@dataclass(frozen=True, slots=True)
class FromCsvText:
    text: str


@dataclass(frozen=True, slots=True)
class FromCsvPath:
    path: Path


@dataclass(frozen=True, slots=True)
class FromExcelPath:
    path: Path
    sheet: int | str = 0


@dataclass(frozen=True, slots=True)
class FromDictRows:
    rows: tuple[Mapping[str, Any], ...]


@dataclass(frozen=True, slots=True)
class FromPackageResource:
    """Read UTF-8 text from ``importlib.resources`` (e.g. package data in site-packages / wheel)."""

    package: str
    resource: str


TableLoadSource: TypeAlias = (
    FromCsvText | FromCsvPath | FromExcelPath | FromDictRows | FromPackageResource
)


def from_csv_text(text: str) -> FromCsvText:
    return FromCsvText(text)


def from_csv_path(path: str | Path) -> FromCsvPath:
    return FromCsvPath(Path(path))


def from_excel_path(path: str | Path, sheet: int | str = 0) -> FromExcelPath:
    return FromExcelPath(Path(path), sheet)


def from_dict_rows(rows: Iterable[Mapping[str, Any]]) -> FromDictRows:
    return FromDictRows(tuple(rows))


def from_package_resource(package: str, resource: str) -> FromPackageResource:
    return FromPackageResource(package, resource)


def _read_csv_matrix(text: str) -> tuple[list[str], list[list[str]]]:
    buf = text.lstrip("\ufeff")
    sample = buf[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        reader = csv.reader(io.StringIO(buf), dialect)
    except csv.Error:
        reader = csv.reader(io.StringIO(buf))
    rows = list(reader)
    if not rows:
        return [], []
    header = [c.strip() for c in rows[0]]
    body: list[list[str]] = []
    for parts in rows[1:]:
        row = [parts[i].strip() if i < len(parts) else "" for i in range(len(header))]
        if any(x.strip() for x in row):
            body.append(row)
    return header, body


def _read_excel_matrix(path: Path, sheet: int | str) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("Excel loading requires openpyxl: pip install openpyxl") from e
    wb = load_workbook(path, read_only=True, data_only=True)
    if isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[str(sheet)]
    it = ws.iter_rows(values_only=True)
    try:
        first = next(it)
    except StopIteration:
        wb.close()
        return [], []
    header = ["" if c is None else str(c).strip() for c in first]
    body: list[list[str]] = []
    for row in it:
        cells = ["" if c is None else str(c).strip() for c in row]
        while len(cells) < len(header):
            cells.append("")
        cells = cells[: len(header)]
        if any(x.strip() for x in cells):
            body.append(cells)
    wb.close()
    return header, body


def _slug_row(header: list[str], cells: list[str]) -> dict[str, str]:
    keys = [slugify_column_name(h) for h in header]
    return {keys[i]: cells[i].strip() if i < len(cells) else "" for i in range(len(keys))}


def _col_index(header: list[str], label: str) -> int | None:
    want = slugify_column_name(label)
    keys = [slugify_column_name(h) for h in header]
    try:
        return keys.index(want)
    except ValueError:
        return None


def _canon_from_header_row(header: list[str], cells: list[str], mapping: ChangeRowMapping) -> dict[str, str]:
    out: dict[str, str] = {}

    def put(label: str, canon: str) -> None:
        idx = _col_index(header, label)
        if idx is None:
            raise KeyError(f"column not found for {canon!r}: {label!r} (slug {slugify_column_name(label)!r})")
        out[canon] = cells[idx].strip() if idx < len(cells) else ""

    put(mapping.change_kind, "change_type")
    put(mapping.source, "source")
    put(mapping.destination, "destination")
    put(mapping.service, "service")
    if mapping.ticket_number:
        put(mapping.ticket_number, "ticket_number")
    if mapping.inf_number:
        put(mapping.inf_number, "inf_number")
    return out


def _canon_from_slug_row(slug_row: Mapping[str, str], mapping: ChangeRowMapping) -> dict[str, str]:
    """Row keys are already slugified (e.g. from ``from_dict_rows`` after ``inputs._row_str``)."""

    def pick(label: str) -> str:
        sk = slugify_column_name(label)
        if sk not in slug_row:
            raise KeyError(f"row missing slug key {sk!r} for label {label!r}")
        return str(slug_row[sk]).strip()

    out: dict[str, str] = {
        "change_type": pick(mapping.change_kind),
        "source": pick(mapping.source),
        "destination": pick(mapping.destination),
        "service": pick(mapping.service),
    }
    if mapping.ticket_number:
        out["ticket_number"] = pick(mapping.ticket_number)
    if mapping.inf_number:
        out["inf_number"] = pick(mapping.inf_number)
    return out


def load_change_rows(source: TableLoadSource, mapping: ChangeRowMapping) -> list[dict[str, str]]:
    if isinstance(source, FromCsvText):
        header, body = _read_csv_matrix(source.text)
    elif isinstance(source, FromCsvPath):
        header, body = _read_csv_matrix(source.path.read_text(encoding="utf-8"))
    elif isinstance(source, FromExcelPath):
        header, body = _read_excel_matrix(source.path, source.sheet)
    elif isinstance(source, FromPackageResource):
        root = resources.files(source.package)
        tr = root
        for seg in Path(source.resource).parts:
            tr = tr / seg
        text = tr.read_text(encoding="utf-8")
        header, body = _read_csv_matrix(text)
    elif isinstance(source, FromDictRows):
        out: list[dict[str, str]] = []
        for raw in source.rows:
            slug_row = {slugify_column_name(str(k)): ("" if v is None else str(v)).strip() for k, v in raw.items()}
            out.append(_canon_from_slug_row(slug_row, mapping))
        return out
    else:
        raise TypeError(source)

    if not header:
        return []
    return [_canon_from_header_row(header, row, mapping) for row in body]
